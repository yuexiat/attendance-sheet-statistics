import openpyxl


def summarize_data():
    wb = openpyxl.load_workbook(r'D:\工程6月.xlsx')  # 替换为实际的Excel文件路径
    ws = wb['总表']  # 将 "Sheet1" 替换为实际的工作表名称

    last_row = ws.max_row
    # print(ws.cell(5,2).value)
    #j是行 ，从第5行开始
    for j in range(5, last_row + 1):
        samename = True
        if ws.cell(j, 2).value is not None:

            #i是列
            for i in range(5,17,3):
                if ws.cell(j, i).value != ws.cell(j, 2).value:

                    samename = False
                    print('不一致',ws.cell(j, i).value)
                    break
                else:
                    print(ws.cell(j, i).value)
            if samename:
                ws.cell(row=j, column=20).value = ws.cell(j, 2).value
                ws.cell(row=j, column=21).value = sum([ws.cell(row=j, column=col).value or 0 for col in range(3, 18, 3)])
                ws.cell(row=j, column=22).value = sum([ws.cell(row=j, column=col).value or 0 for col in range(4, 19, 3)])
    # for i in range(5, last_row + 1):
    #     name = ws.cell(row=i, column=2).value
    #     value1 = sum([ws.cell(row=i, column=col).value or 0 for col in range(3, 9, 3)])
    #     value2 = sum([ws.cell(row=i, column=col).value or 0 for col in range(4, 10, 3)])
    #
    #     if name == ws.cell(row=i - 1, column=2).value:
    #         ws.cell(row=i, column=20).value = name
    #         ws.cell(row=i, column=21).value = (ws.cell(row=i - 1, column=21).value or 0) + value1
    #         ws.cell(row=i, column=22).value = (ws.cell(row=i - 1, column=22).value or 0) + value2
    #     else:
    #         ws.cell(row=i, column=20).value = name
    #         ws.cell(row=i, column=21).value = value1
    #         ws.cell(row=i, column=22).value = value2

    wb.save(r'D:\output_file.xlsx')  # 替换为输出结果的Excel文件路径
    wb.close()


summarize_data()