import openpyxl
from openpyxl.styles import Alignment


def summarize_data():
    wb = openpyxl.load_workbook(r'D:\Work file\表格\工程6月.xlsx')  # 替换为实际的Excel文件路径
    ws = wb['总表']  # 将 "Sheet1" 替换为实际的工作表名称
    wb_hz = openpyxl.load_workbook(r'D:\Work file\表格\1-6汇总.xlsx')  # 替换为实际的Excel文件路径
    ws_hz = wb_hz['Sheet1']  # 将 "Sheet1" 替换为实际的工作表名称
    last_row = ws.max_row

    for j in range(5, last_row + 1):
        samename = True
        if ws.cell(j, 2).value is not None:
            name = ws.cell(row=j, column=2).value
            attendance = ws.cell(row=j, column=3).value
            actual_time = ws.cell(row=j, column=4).value
            late = ws.cell(row=j, column=5).value
            early_leave = ws.cell(row=j, column=6).value
            no_punch = ws.cell(row=j, column=7).value

            ws_overtime = wb['加班']
            row_number = find_row_by_name(ws_overtime, 2, name)
            overtime = 0
            if row_number is not None:
                overtime = ws_overtime.cell(row=row_number, column=34).value #加班总数
            total_hours = (actual_time or 0) + (overtime or 0)
            late_count = late or 0
            early_leave_count = early_leave or 0
            no_punch_count = no_punch or 0

            ws_hz.cell(j-3,2).value = name

            ws_hz.cell(j - 3, 8).value = '应出勤:'+str(attendance)+'\n实际出勤:'+str(actual_time)+'\n加班时数:'+str(overtime)+'\n总时数:'+str(total_hours)+'\n迟到:'+str(late_count)+'\n早退:'+str(early_leave_count)+'\n无打卡次数:'+str(no_punch_count)
            ws_hz.cell(j - 3, 8).alignment = Alignment(wrapText=True)

    wb.close()
    wb_hz.save(r'D:\output_file.xlsx')  # 替换为输出结果的Excel文件路径
    wb_hz.close()

def summarize_data1():
    wb = openpyxl.load_workbook(r'D:\Work file\表格\工程1月.xlsx')  # 替换为实际的Excel文件路径
    ws = wb['总表']  # 将 "Sheet1" 替换为实际的工作表名称
    wb_hz = openpyxl.load_workbook(r'D:\output_file.xlsx')  # 替换为实际的Excel文件路径
    ws_hz = wb_hz['Sheet1']  # 将 "Sheet1" 替换为实际的工作表名称
    last_row = ws_hz.max_row

    for j in range(2, last_row + 1):
        samename = True
        if ws_hz.cell(j, 2).value is not None:

            name = ws_hz.cell(row=j, column=2).value
            print(name,1)
            row_number_name = find_row_by_name(ws, 2, name) #以模版姓名为基准
            if row_number_name is None:
                continue
            attendance = ws.cell(row=row_number_name, column=3).value
            actual_time = ws.cell(row=row_number_name, column=4).value
            late = ws.cell(row=row_number_name, column=5).value
            early_leave = ws.cell(row=row_number_name, column=6).value
            no_punch = ws.cell(row=row_number_name, column=7).value

            ws_overtime = wb['加班']
            row_number = find_row_by_name(ws_overtime, 2, name)
            overtime = 0
            if row_number is not None:
                overtime = ws_overtime.cell(row=row_number, column=34).value #加班总数
            total_hours = float(actual_time or 0) + float(overtime or 0)

            late_count = late or 0
            early_leave_count = early_leave or 0
            no_punch_count = no_punch or 0

            ws_hz.cell(j,2).value = name

            ws_hz.cell(j, 3).value = '应出勤:'+str(attendance)+'\n实际出勤:'+str(actual_time)+'\n加班时数:'+str(overtime)+'\n总时数:'+str(total_hours)+'\n迟到:'+str(late_count)+'\n早退:'+str(early_leave_count)+'\n无打卡次数:'+str(no_punch_count)
            ws_hz.cell(j, 3).alignment = Alignment(wrapText=True)

    wb.close()
    wb_hz.save(r'D:\output_file.xlsx')  # 替换为输出结果的Excel文件路径
    wb_hz.close()

def find_row_by_name(sheet, column_index, name):
    for row in range(1, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=column_index).value
        if cell_value == name:
            return row
    return None

summarize_data1()