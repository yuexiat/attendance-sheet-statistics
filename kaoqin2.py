import openpyxl
from openpyxl.styles import Alignment


def summarize_data(n):
    wb = openpyxl.load_workbook(r'D:\Work file\表格\工程'+str(n)+'月.xlsx')  # 替换为实际的Excel文件路径
    ws = wb['总表']  # 将 "Sheet1" 替换为实际的工作表名称
    wb_hz = openpyxl.load_workbook(r'.\考勤模板.xlsx')  # 替换为实际的Excel文件路径
    ws_hz = wb_hz['Sheet1']  # 将 "Sheet1" 替换为实际的工作表名称
    last_row = ws.max_row

    for j in range(5, last_row + 1):
        samename = True
        if ws.cell(j, 2).value is not None:
            name = ws.cell(row=j, column=2).value
            attendance = ws.cell(row=j, column=3).value
            actual_time = ws.cell(row=j, column=4).value
            late = ws.cell(row=j, column=5).value
            early_leave = ws.cell(row=j, column=6).value
            no_punch = ws.cell(row=j, column=7).value

            ws_overtime = wb['加班表']
            row_number = find_row_by_name(ws_overtime, 2, name)
            c_number = find_column_by_name(ws_overtime, 3, "累计加班合计")
            overtime = 0
            if row_number is not None:
                overtime = ws_overtime.cell(row=row_number, column=c_number).value #加班总数
            total_hours = (actual_time or 0) + (overtime or 0)
            late_count = late or 0
            early_leave_count = early_leave or 0
            no_punch_count = no_punch or 0

            ws_hz.cell(2+(j-5)*7,2).value = name


            column1 = n+3

            ws_hz.cell(2+(j-5)*7, column1).value = attendance
            ws_hz.cell(2+(j-5)*7, column1).alignment = Alignment(wrapText=True)

            ws_hz.cell(2 + (j - 5) * 7 + 1, column1).value = actual_time
            ws_hz.cell(2 + (j - 5) * 7 + 1, column1).alignment = Alignment(wrapText=True)
            ws_hz.cell(2 + (j - 5) * 7 + 2, column1).value = overtime
            ws_hz.cell(2 + (j - 5) * 7 + 2, column1).alignment = Alignment(wrapText=True)
            ws_hz.cell(2 + (j - 5) * 7 + 3, column1).value = total_hours
            ws_hz.cell(2 + (j - 5) * 7 + 3, column1).alignment = Alignment(wrapText=True)
            ws_hz.cell(2 + (j - 5) * 7 + 4, column1).value = late_count
            ws_hz.cell(2 + (j - 5) * 7 + 4, column1).alignment = Alignment(wrapText=True)
            ws_hz.cell(2 + (j - 5) * 7 + 5, column1).value = early_leave_count
            ws_hz.cell(2 + (j - 5) * 7 + 5, column1).alignment = Alignment(wrapText=True)
            ws_hz.cell(2 + (j - 5) * 7 + 6, column1).value = no_punch_count
            ws_hz.cell(2 + (j - 5) * 7 + 6, column1).alignment = Alignment(wrapText=True)
            #
            # for s in range(1,7):
            #     ws_hz.cell(2 + (j - 5) * 7 + s, column1).number_format = numbers.FORMAT_NUMBER

            # ws_hz.cell(j - 3, 8).alignment = Alignment(wrapText=True)

    wb.close()
    wb_hz.save(r'D:\output_file.xlsx')  # 替换为输出结果的Excel文件路径
    wb_hz.close()

def summarize_data1(n):
    wb = openpyxl.load_workbook(r'D:\Work file\表格\工程'+str(n)+'月.xlsx')  # 替换为实际的Excel文件路径
    ws = wb['总表']  # 将 "Sheet1" 替换为实际的工作表名称
    wb_hz = openpyxl.load_workbook(r'D:\output_file.xlsx')  # 替换为实际的Excel文件路径
    ws_hz = wb_hz['Sheet1']  # 将 "Sheet1" 替换为实际的工作表名称
    last_row = ws_hz.max_row

    for j in range(2, last_row + 1):
        samename = True
        if ws_hz.cell(j, 2).value is not None:

            name = ws_hz.cell(row=j, column=2).value
            # print(name,1)
            row_number_name = find_row_by_name(ws, 2, name) #以模版姓名为基准
            if row_number_name is None:
                continue
            attendance = ws.cell(row=row_number_name, column=3).value
            actual_time = ws.cell(row=row_number_name, column=4).value
            late = ws.cell(row=row_number_name, column=5).value
            early_leave = ws.cell(row=row_number_name, column=6).value
            no_punch = ws.cell(row=row_number_name, column=7).value

            ws_overtime = wb['加班表']
            row_number = find_row_by_name(ws_overtime, 2, name)
            overtime = 0
            c_number =find_column_by_name(ws_overtime, 3, "累计加班合计")
            if row_number is not None:
                overtime = ws_overtime.cell(row=row_number, column=c_number).value #加班总数
            total_hours = float(actual_time or 0) + float(overtime or 0)

            late_count = late or 0
            early_leave_count = early_leave or 0
            no_punch_count = no_punch or 0

            ws_hz.cell(j,2).value = name


            # ws_hz.cell(j, 3).value = '应出勤:'+str(attendance)+'\n实际出勤:'+str(actual_time)+'\n加班时数:'+str(overtime)+'\n总时数:'+str(total_hours)+'\n迟到:'+str(late_count)+'\n早退:'+str(early_leave_count)+'\n无打卡次数:'+str(no_punch_count)


            column1 = n+3

            ws_hz.cell(j, column1).value = attendance
            ws_hz.cell(j, column1).alignment = Alignment(wrapText=True)
            ws_hz.cell(j + 1, column1).value = actual_time
            ws_hz.cell(j + 1, column1).alignment = Alignment(wrapText=True)
            ws_hz.cell(j + 2, column1).value = overtime
            ws_hz.cell(j + 2, column1).alignment = Alignment(wrapText=True)
            ws_hz.cell(j + 3, column1).value = total_hours
            ws_hz.cell(j + 3, column1).alignment = Alignment(wrapText=True)
            ws_hz.cell(j + 4, column1).value = late_count
            ws_hz.cell(j + 4, column1).alignment = Alignment(wrapText=True)
            ws_hz.cell(j + 5, column1).value = early_leave_count
            ws_hz.cell(j + 5, column1).alignment = Alignment(wrapText=True)
            ws_hz.cell(j + 6, column1).value = no_punch_count
            ws_hz.cell(j + 6, column1).alignment = Alignment(wrapText=True)
            attendance = 0
            actual_time = 0
            overtime = 0
            total_hours = 0
            late_count = 0
            early_leave_count = 0
            no_punch_count = 0

            # for s in range(0,7):
            #     ws_hz.cell(j + s, column1).number_format = numbers.FORMAT_NUMBER
    wb.close()
    wb_hz.save(r'D:\output_file.xlsx')  # 替换为输出结果的Excel文件路径
    wb_hz.close()


def find_row_by_name(sheet, column_index, name):
    for row in range(1, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=column_index).value
        if cell_value == name:
            print(name,row)
            return row

    return None

def find_column_by_name(sheet, row, name):
    # Iterate through the specified column to find the target name
    for column_index in range(1, 99):
        cell_value = sheet.cell(row=row, column=column_index).value
        if cell_value == name:
            return column_index  # Return the row index
    return None  # Return None if the name is not found


def append_name():
    wb_hz = openpyxl.load_workbook(r'D:\output_file.xlsx')  # 替换为实际的Excel文件路径
    ws_hz = wb_hz['Sheet1']  # 将 "Sheet1" 替换为实际的工作表名称
    last_row = ws_hz.max_row

    wb1_hz = openpyxl.load_workbook(r'.\人员和组.xlsx')  # 替换为实际的Excel文件路径
    ws1_hz = wb1_hz['Sheet1']  # 将 "Sheet1" 替换为实际的工作表名称

    for row_index in range(2,last_row+1,7): #第二行开始，每间隔7行
        name = ws_hz.cell(row=row_index, column=2).value  #初始赋值名字
        zb = find_row_by_name(ws1_hz, 2, name)  #查找组所在的行

        if zb is not None: #把查到的组别赋值到output里
            ws_hz.cell(row_index, 1).value = ws1_hz.cell(zb, 1).value

        if name is not None:
            for row_name in range(1,7):#重复6次，1-6
                ws_hz.cell(row_index+row_name, 2).value = name
                if zb is not None:#把查到的组别赋值到output里
                    ws_hz.cell(row_index+row_name, 1).value = ws1_hz.cell(zb, 1).value






    wb_hz.save(r'D:\output_file.xlsx')  # 替换为输出结果的Excel文件路径
    wb_hz.close()

if __name__ == '__main__':
    summarize_data(9)#最后一个月的月份
    for i in range(8,6,-1):
        summarize_data1(i)
    append_name()
